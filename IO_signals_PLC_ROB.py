import csv
import openpyxl as xl
from pathlib import Path


def get_robot_plc_signals(sheet, signals_with_descriptions={}):
    '''
    Extracts plc signals with descriptions from single xlsm sheet and, returns updated signals with descriptions.


    Parameters:
        sheet (Worksheet): The single sheet from which plc signals and descriptions are to be extracted.
        signals_with_descriptions (dict): Signals with descriptions which are to be updated.


    Returns:
        signals_with_descriptions (dict): Signals and descriptions which get updated with plc signals.
    '''

    start_plc_signals_row = 7
    end_plc_signals_row = 31

    signal_number_column = 11

    start_outputs_column = 12
    end_outputs_column = 15
    start_inputs_colum = 16
    end_inputs_column = 19

    for row in range(start_plc_signals_row, end_plc_signals_row):

        signal_number = sheet.cell(row, signal_number_column).value
        output_signal_description = ""
        input_signal_description = ""

        for column in range(start_outputs_column, end_outputs_column + 1):

            actual_cell_value = sheet.cell(row, column).value

            if actual_cell_value:
                output_signal_description += str(actual_cell_value) + " "
            else:
                break

        for column in range(start_inputs_colum, end_inputs_column + 1):

            actual_cell_value = sheet.cell(row, column).value

            if actual_cell_value:
                input_signal_description += str(actual_cell_value) + " "
            else:
                break

        if(output_signal_description != ""):

            signal_with_description = {
                "A"+str(signal_number): output_signal_description
            }

            signals_with_descriptions.update(signal_with_description)

        if(input_signal_description != ""):

            signal_with_description = {
                "E"+str(signal_number): input_signal_description
            }

            signals_with_descriptions.update(signal_with_description)

    return signals_with_descriptions


def get_robot_collisions_signals(sheet, signals_with_descriptions={}):
    '''
    Extracts collisions signals with descriptions from single xlsm sheet and, returns updated signals with descriptions.


    Parameters:
        sheet (Worksheet): The single sheet from which collisions signals and descriptions are to be extracted.
        signals_with_descriptions (dict): Signals with descriptions which are to be updated.


    Returns:
        signals_with_descriptions (dict): Signals and descriptions which get updated with collisions signals.
    '''

    start_collisions_signals_row = 7
    end_collisions_signals_row = 22

    signal_column_number = 22

    start_collisions_signals_column = 23
    end_collisions_signals_column = 38

    for column in range(start_collisions_signals_column, end_collisions_signals_column + 1):
        robot_name = sheet.cell(start_collisions_signals_row-1, column).value
        if robot_name:
            for row in range(start_collisions_signals_row, end_collisions_signals_row + 1):
                if sheet.cell(row, column).value == "X":
                    collision_signal = sheet.cell(
                        row, signal_column_number).value

                    signals_with_descriptions.update(
                        {
                            f"E{collision_signal}": f"Robotzone {int(collision_signal)-40} free Rob < {robot_name}",
                            f"E{int(collision_signal)+40}": f"Acknowledge robotcollision {int(collision_signal)-40} Rob > {robot_name}",
                            f"A{int(collision_signal)+40}": f"Request robotcollision {int(collision_signal)-40} Rob > {robot_name}",
                            f"A{collision_signal}": f"Release robotcollision {int(collision_signal)-40} Rob > {robot_name}",

                        }
                    )

    return signals_with_descriptions
